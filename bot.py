import os
import pandas as pd

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder,
    MessageHandler,
    CommandHandler,
    CallbackQueryHandler,
    filters,
    ContextTypes
)

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

TOKEN = os.getenv("BOT_TOKEN")

user_files = {}
user_colors = {}


# ---------- COLOR HELPERS ----------

def hex_to_rgb(hex_color):
    hex_color = hex_color.replace("#","")
    return tuple(int(hex_color[i:i+2],16) for i in (0,2,4))

def rgb_to_hex(rgb):
    return "%02x%02x%02x" % rgb

def lighten(hex_color, factor=0.2):

    r,g,b = hex_to_rgb(hex_color)

    r = int(r + (255-r)*factor)
    g = int(g + (255-g)*factor)
    b = int(b + (255-b)*factor)

    return rgb_to_hex((r,g,b))

def darken(hex_color, factor=0.2):

    r,g,b = hex_to_rgb(hex_color)

    r = int(r*(1-factor))
    g = int(g*(1-factor))
    b = int(b*(1-factor))

    return rgb_to_hex((r,g,b))


def is_dark(hex_color):

    r,g,b = hex_to_rgb(hex_color)

    brightness = (r*299 + g*587 + b*114)/1000

    return brightness < 140


# ---------- START ----------

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):

    text = """
🎨 *Welcome to FIXCEL*

Send an Excel file and I will format it beautifully.

Features
• Custom header colors
• Alternating rows
• Wrap text
• Clean borders
• Smart alignment

Free Plan
2 files per month

Premium
Unlimited formatting
$6/month
"""

    await update.message.reply_text(text, parse_mode="Markdown")


# ---------- FILE UPLOAD ----------

async def handle_file(update: Update, context: ContextTypes.DEFAULT_TYPE):

    user = update.message.from_user.id
    file = update.message.document

    tg_file = await file.get_file()

    path = f"{user}_input.xlsx"

    await tg_file.download_to_drive(path)

    user_files[user] = path

    keyboard = [

        [InlineKeyboardButton("Use Excel Brand Color (#1D6F42)", callback_data="excel_color")],
        [InlineKeyboardButton("Use Custom HEX Color", callback_data="custom_color")]

    ]

    await update.message.reply_text(

        "🎨 Choose header color style:",
        reply_markup=InlineKeyboardMarkup(keyboard)

    )


# ---------- COLOR CHOICE ----------

async def color_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):

    query = update.callback_query
    await query.answer()

    user = query.from_user.id

    if query.data == "excel_color":

        result = format_excel(user_files[user], "#1D6F42", 1)

        await query.message.reply_document(document=open(result,"rb"))

        return

    await query.message.reply_text(
        "Send preferred HEX color.\nExample:\n#FF5733"
    )


# ---------- RECEIVE HEX ----------

async def receive_hex(update: Update, context: ContextTypes.DEFAULT_TYPE):

    user = update.message.from_user.id

    if user not in user_files:
        return

    hex_color = update.message.text.strip()

    if not hex_color.startswith("#") or len(hex_color)!=7:

        await update.message.reply_text("Invalid HEX. Example: #FF5733")
        return

    user_colors[user] = hex_color

    keyboard = [

        [
            InlineKeyboardButton("1 Color", callback_data="p1"),
            InlineKeyboardButton("2 Colors", callback_data="p2")
        ],
        [
            InlineKeyboardButton("3 Colors", callback_data="p3"),
            InlineKeyboardButton("4 Colors", callback_data="p4")
        ]

    ]

    await update.message.reply_text(

        "How many header colors should be used?",
        reply_markup=InlineKeyboardMarkup(keyboard)

    )


# ---------- PATTERN CHOICE ----------

async def pattern_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):

    query = update.callback_query
    await query.answer()

    user = query.from_user.id

    pattern = int(query.data[1])

    color = user_colors[user]

    result = format_excel(user_files[user], color, pattern)

    await query.message.reply_document(document=open(result,"rb"))


# ---------- EXCEL FORMAT ----------

def format_excel(file, base_color, pattern):

    df = pd.read_excel(file)

    output = "formatted.xlsx"

    df.to_excel(output, index=False)

    wb = load_workbook(output)

    ws = wb.active

    ws.insert_rows(1)
    ws.insert_cols(1)

    thin = Side(style="thin", color="b7b7b7")

    border = Border(left=thin,right=thin,top=thin,bottom=thin)

    gray_fill = PatternFill(start_color="F2F2F2", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", fill_type="solid")

    max_row = ws.max_row
    max_col = ws.max_column


    # -------- GENERATE HEADER COLORS --------

    colors = [base_color.replace("#","")]

    if pattern>=2:
        colors.append(lighten(base_color,0.2))

    if pattern>=3:
        colors.append(darken(base_color,0.2))

    if pattern>=4:
        colors.append(lighten(base_color,0.4))


    # if more columns than colors generate shades

    while len(colors)<max_col:

        colors.append(lighten("#"+colors[-1],0.15))


    font_color = "FFFFFF" if is_dark(base_color) else "333333"


    # -------- HEADER STYLE --------

    for c in range(2,max_col+1):

        color = colors[(c-2)%pattern]

        cell = ws.cell(row=2,column=c)

        cell.fill = PatternFill(start_color=color, fill_type="solid")

        cell.font = Font(color=font_color,bold=True)

        cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)

        cell.border = border


    # -------- DATA CELLS --------

    for r in range(3,max_row+1):

        for c in range(2,max_col+1):

            cell = ws.cell(row=r,column=c)

            cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)

            cell.border = border

            if r%2==1:
                cell.fill = gray_fill
            else:
                cell.fill = white_fill


    # -------- CLEAN BACKGROUND --------

    for r in range(1, ws.max_row+50):

        for c in range(1, ws.max_column+50):

            if r>=2 and c>=2 and r<=max_row and c<=max_col:
                continue

            cell = ws.cell(row=r,column=c)

            cell.fill = white_fill
            cell.border = Border()


    ws.column_dimensions['A'].width = 2


    wb.save(output)

    return output


# ---------- BOT SETUP ----------

app = ApplicationBuilder().token(TOKEN).build()

app.add_handler(CommandHandler("start", start))

app.add_handler(MessageHandler(filters.Document.ALL, handle_file))

app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, receive_hex))

app.add_handler(CallbackQueryHandler(color_choice, pattern="excel_color|custom_color"))

app.add_handler(CallbackQueryHandler(pattern_choice, pattern="p[1-4]"))

app.run_polling()
